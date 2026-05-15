"""Statutory schedule exporters — CSV, XLSX, PDF.

Used by the May 2026 audit requirement that PAYE / NSSF / SHIF templates be
exportable in Excel, CSV and PDF for submission to relevant government portals.

All exporters return ``bytes``. ReportLab and openpyxl are imported lazily so
the rest of the system still works when those packages are unavailable.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from typing import Any, Literal

_MONTHS = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

ScheduleKind = Literal["paye", "nssf", "shif"]


# ── Header definitions (column order matters for KRA/NSSF/SHIF portals) ──


def _paye_headers() -> list[str]:
    return [
        "Employee",
        "KRA PIN",
        "Gross Pay",
        "Taxable Pay",
        "PAYE Tax",
        "Personal Relief",
        "Net Tax",
    ]


def _nssf_headers() -> list[str]:
    return [
        "Employee",
        "NSSF Number",
        "Pensionable Pay",
        "Tier I",
        "Tier II",
        "Total",
    ]


def _shif_headers() -> list[str]:
    return [
        "Employee",
        "SHIF Number",
        "Gross Pay",
        "SHIF Amount",
    ]


# NSSF tier limits (2024 Act). Tier I = first 8,000; Tier II = 8,001 – 72,000.
# Total employee rate is 6% capped per tier.
_NSSF_TIER1_CAP = Decimal("8000")
_NSSF_RATE = Decimal("0.06")
_KE_PERSONAL_RELIEF = Decimal("2400")


def _split_paye_row(row: dict[str, Any]) -> list[str]:
    """Map an internal PAYE row to the KRA P10 column layout."""
    gross = Decimal(row.get("gross") or 0)
    amount = Decimal(row.get("amount") or 0)
    relief = _KE_PERSONAL_RELIEF if amount > 0 else Decimal("0")
    net_tax = amount  # service already nets out relief in `paye`
    return [
        str(row.get("full_name") or ""),
        str(row.get("id_value") or row.get("kra_pin") or ""),
        f"{gross:.2f}",
        f"{gross:.2f}",
        f"{(amount + relief):.2f}",
        f"{relief:.2f}",
        f"{net_tax:.2f}",
    ]


def _split_nssf_row(row: dict[str, Any]) -> list[str]:
    """Map an NSSF row into tier I / tier II / total."""
    pensionable = Decimal(row.get("gross") or 0)
    total = Decimal(row.get("amount") or 0)
    tier_one = (min(pensionable, _NSSF_TIER1_CAP) * _NSSF_RATE).quantize(Decimal("0.01"))
    tier_two = (total - tier_one).quantize(Decimal("0.01"))
    if tier_two < 0:
        tier_two = Decimal("0.00")
    return [
        str(row.get("full_name") or ""),
        str(row.get("id_value") or ""),
        f"{pensionable:.2f}",
        f"{tier_one:.2f}",
        f"{tier_two:.2f}",
        f"{total:.2f}",
    ]


def _split_shif_row(row: dict[str, Any]) -> list[str]:
    return [
        str(row.get("full_name") or ""),
        str(row.get("id_value") or ""),
        f"{Decimal(row.get('gross') or 0):.2f}",
        f"{Decimal(row.get('amount') or 0):.2f}",
    ]


_HEADER_FOR: dict[ScheduleKind, list[str]] = {
    "paye": _paye_headers(),
    "nssf": _nssf_headers(),
    "shif": _shif_headers(),
}


def _row_mapper(kind: ScheduleKind):
    return {
        "paye": _split_paye_row,
        "nssf": _split_nssf_row,
        "shif": _split_shif_row,
    }[kind]


# ── CSV ────────────────────────────────────────────────────────────────────


def export_schedule_csv(
    kind: ScheduleKind,
    rows: list[dict[str, Any]],
    month: int,
    year: int,
) -> bytes:
    """Render a statutory schedule as CSV bytes (UTF-8).

    @param kind: "paye" | "nssf" | "shif"
    @param rows: Per-employee schedule rows (output of get_*_schedule)
    @param month: Period month
    @param year: Period year
    @returns CSV file content as bytes
    """
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow([f"{kind.upper()} Schedule — {_MONTHS[month - 1]} {year}"])
    writer.writerow([])
    writer.writerow(_HEADER_FOR[kind])
    mapper = _row_mapper(kind)
    for row in rows:
        writer.writerow(mapper(row))
    return buf.getvalue().encode("utf-8")


# ── XLSX ───────────────────────────────────────────────────────────────────


def export_schedule_xlsx(
    kind: ScheduleKind,
    rows: list[dict[str, Any]],
    month: int,
    year: int,
) -> bytes:
    """Render a statutory schedule as XLSX bytes via openpyxl.

    Falls back to a SpreadsheetML (.xls) document if openpyxl is unavailable —
    the file still opens cleanly in Excel and LibreOffice.

    @returns XLSX bytes (or SpreadsheetML fallback)
    """
    headers = _HEADER_FOR[kind]
    mapper = _row_mapper(kind)
    data_rows = [mapper(r) for r in rows]
    period = f"{_MONTHS[month - 1]} {year}"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = f"{kind.upper()} {period}"

        # Title row
        ws.cell(row=1, column=1, value=f"{kind.upper()} Schedule — {period}").font = Font(
            bold=True, size=14
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        # Header row
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, name in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Body
        for r_idx, row in enumerate(data_rows, start=4):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Column widths — minimal sane defaults
        for col_idx, name in enumerate(headers, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = max(14, len(name) + 4)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        # SpreadsheetML 2003 fallback (no openpyxl dependency)
        def _esc(v: str) -> str:
            return (
                str(v)
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
            )

        header_xml = "<Row>" + "".join(
            f'<Cell><Data ss:Type="String">{_esc(h)}</Data></Cell>' for h in headers
        ) + "</Row>"
        body_xml = "".join(
            "<Row>"
            + "".join(
                f'<Cell><Data ss:Type="String">{_esc(c)}</Data></Cell>' for c in row
            )
            + "</Row>"
            for row in data_rows
        )
        xml = (
            '<?xml version="1.0"?>\n'
            '<?mso-application progid="Excel.Sheet"?>\n'
            '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"\n'
            ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">\n'
            f' <Worksheet ss:Name="{kind.upper()} {period}"><Table>'
            f"{header_xml}{body_xml}</Table></Worksheet>\n"
            "</Workbook>\n"
        )
        return xml.encode("utf-8")


# ── PDF ────────────────────────────────────────────────────────────────────


def export_schedule_pdf(
    kind: ScheduleKind,
    rows: list[dict[str, Any]],
    month: int,
    year: int,
) -> bytes:
    """Render a statutory schedule as a landscape A4 PDF via ReportLab.

    Falls back to plain text if ReportLab is unavailable.
    """
    headers = _HEADER_FOR[kind]
    mapper = _row_mapper(kind)
    data_rows = [mapper(r) for r in rows]
    period = f"{_MONTHS[month - 1]} {year}"

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=14 * mm,
            bottomMargin=14 * mm,
            title=f"{kind.upper()} Schedule {year}-{month:02d}",
        )
        styles = getSampleStyleSheet()
        story: list[Any] = []
        story.append(Paragraph(f"<b>{kind.upper()} Schedule — {period}</b>", styles["Heading2"]))
        story.append(Spacer(1, 6))

        # Body table
        table_data: list[list[str]] = [headers, *data_rows]
        if len(data_rows) == 0:
            table_data.append(["(no rows)"] + [""] * (len(headers) - 1))

        tbl = Table(table_data, repeatRows=1)
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ]
            )
        )
        story.append(tbl)
        story.append(Spacer(1, 8))
        story.append(
            Paragraph(
                f"<font size='8' color='grey'>Powered by Aifya · {len(data_rows)} employee(s)</font>",
                styles["Normal"],
            )
        )
        doc.build(story)
        return buf.getvalue()
    except ImportError:
        # Plain text fallback
        lines = [f"{kind.upper()} Schedule — {period}", ""]
        lines.append(" | ".join(headers))
        for row in data_rows:
            lines.append(" | ".join(row))
        return "\n".join(lines).encode("utf-8")


# ── Top-level convenience dispatch ─────────────────────────────────────────


def render_schedule(
    kind: ScheduleKind,
    fmt: str,
    rows: list[dict[str, Any]],
    month: int,
    year: int,
) -> tuple[bytes, str, str]:
    """Render a schedule in the requested format.

    @param kind: "paye" | "nssf" | "shif"
    @param fmt: "csv" | "xlsx" | "pdf"
    @param rows: Schedule rows
    @param month: Period month
    @param year: Period year
    @returns Tuple of (bytes, media_type, filename)
    @raises ValueError: If fmt is not one of csv|xlsx|pdf
    """
    fmt_lower = fmt.lower()
    base = f"{kind}-schedule-{year}-{month:02d}"
    if fmt_lower == "csv":
        return export_schedule_csv(kind, rows, month, year), "text/csv", f"{base}.csv"
    if fmt_lower == "xlsx":
        data = export_schedule_xlsx(kind, rows, month, year)
        # When openpyxl is missing we emit SpreadsheetML — surface the .xls extension
        is_xlsx = data[:4] == b"PK\x03\x04"
        return (
            data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if is_xlsx
            else "application/vnd.ms-excel",
            f"{base}.{'xlsx' if is_xlsx else 'xls'}",
        )
    if fmt_lower == "pdf":
        data = export_schedule_pdf(kind, rows, month, year)
        return (
            data,
            "application/pdf" if data[:4] == b"%PDF" else "text/plain",
            f"{base}.pdf",
        )
    raise ValueError(f"Unsupported format: {fmt}. Use csv | xlsx | pdf.")
